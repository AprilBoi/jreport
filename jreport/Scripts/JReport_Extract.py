from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from datetime import datetime,timedelta,date
import datetime as dt
import time
import os
import os.path
import pandas as pd
import openpyxl



class JReportExtract:
    def __init__(self, driver_path: str):
        self.driver_path = driver_path
        
        self.driver = None

    def start_driver(self):          
        
        service = ChromeService(executable_path=self.driver_path)
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")                  
        self.driver = webdriver.Chrome(service=service, options=options)      
        
    def download_data(self, url: str, report_date):
        
        if not self.driver:
            print('start')
            self.start_driver()        

        self.driver.get(url)           
        
        print(f'{reportdate=}')

        report_startdate = self.driver.find_element("id", "sdate")
        report_startdate.send_keys(report_date)

        report_enddate = self.driver.find_element("id", "edate")
        report_enddate.send_keys(report_date)

        report_enddate.submit()
        time.sleep(5)       

    def html_xlsx(self,downloaded_file,newfilelocation):        
        df_list = pd.read_html(downloaded_file)
        df = pd.DataFrame(df_list[0])          
        df.to_excel(newfilelocation,index=False)
        
    def data_cleanup(self,queue_path):
                   # Logic from rename
        for filename in os.listdir(queue_path):
           
            if filename.endswith('.xlsx'):
                file_path = os.path.join(queue_path, filename)
                print(f'Update column J for file {file_path=}')                 
                # Load the Excel workbook
                workbook = openpyxl.load_workbook(file_path)
                workbook["Sheet1"].title = "CA COG"
                ws = workbook["CA COG"]
                workbook["CA COG"].delete_rows(ws.min_row, 1)
                for row in range(2, ws.max_row+1):
                    ws["{}{}".format("J", row)].number_format = '@'  
                    
                # Save the changes
                updated_file_path = os.path.join(queue_path, filename)
                workbook.save(updated_file_path)
    
    def stop_driver(self):
        if self.driver:
            self.driver.quit()
            self.driver = None
    


if __name__ == "__main__":

    current_dir = os.getcwd()
    print(current_dir)    
    driver_path = r"jreport\Scripts\chromedriver.exe"    
    os.path.isfile(driver_path)
    print(f'Chrome driver is available {os.path.isfile(driver_path)}')
    # breakpoint()
    queue_path = r'jreport\\Queue\\' 
    # create queue folder if not exist 
    if not os.path.exists(queue_path):
        print('create queue folder')
        os.makedirs(queue_path)
      
    end_date = dt.date.today() - dt.timedelta(days=5)     
     
    start_date = end_date - dt.timedelta(days=0) 
    # print(start_date)
    # breakpoint()
    date_range = [start_date + dt.timedelta(days=delta) for delta in range((end_date - start_date).days + 1)]
    downloader = JReportExtract(driver_path)        
    url = "https://uschizwweb1504/mcmprpt/jReport.asp"           
    
    try: 
        for reportdate in date_range:
            file_report_date = reportdate.strftime('%Y%m%d')             
            downloader.download_data(url, reportdate.strftime('%m-%d-%Y')  )              
            downloaded_file = os.path.join(os.path.expanduser("~"), "Downloads\\results.xls")                                             
            newfilelocation = f'{queue_path}CA COG {file_report_date}.xlsx'
            
            if os.path.exists(downloaded_file):            
                downloader.html_xlsx(downloaded_file,newfilelocation)          
                os.remove(downloaded_file)                     
            else:
                print("wala")  

        downloader.data_cleanup(queue_path)     
        
        
    finally:        
        print('Tapos Na!!!!')
        downloader.stop_driver()
        
        