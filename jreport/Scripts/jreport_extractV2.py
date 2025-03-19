 
import datetime as dt 
import time
import os
 

from selenium.webdriver.common.by import By
from selenium import webdriver 
from selenium.webdriver.chrome.service import Service
from datetime import datetime 
import pandas as pd
import openpyxl
import socket

print(socket.gethostname())
breakpoint()
# create queue folder if not exist 
if socket.gethostname() == 'SGDCIZWAPP1102':
    queue_path = r'D:\SFTPRoot\PROD\JREPORT\Queue'
else:
    queue_path = r'jreport\\Queue\\' 

if not os.path.exists(queue_path):
    print('create queue folder')
    os.makedirs(queue_path)

downloaded_file = os.path.join(os.path.expanduser("~"), "Downloads\\results.xls")   


end_date = dt.date.today() - dt.timedelta(days=1)     
start_date = end_date - dt.timedelta(days=1) 
date_range = [start_date + dt.timedelta(days=delta) for delta in range((end_date - start_date).days + 1)]   
reportdate = end_date 
# for reportdate in date_range:

print ('Executing Job: Freshdesk Call Metrics Download Started at ' + datetime.now().strftime("%Y%m%d-%H:%M:%S"))
dir_path = os.getcwd()
service = Service(r"jreport\Scripts\chromedriver.exe")
options = webdriver.ChromeOptions()
# options.add_argument("start-maximized")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
driver = webdriver.Chrome(service=service, options=options)
driver.get("https://uschizwweb1504/mcmprpt/jReport.asp")

file_report_date = reportdate.strftime('%Y%m%d')   
newfilelocation = f'{queue_path}CA COG {file_report_date}.xlsx'

report_startdate = driver.find_element(By.ID, "sdate")
report_startdate.send_keys(reportdate.strftime('%m/%d/%Y'))

report_enddate = driver.find_element(By.ID, "edate")
report_enddate.send_keys(reportdate.strftime('%m/%d/%Y'))
report_enddate.submit() 

time.sleep(3)
downloaded_file = os.path.join(os.path.expanduser("~"), "Downloads\\results.xls")                                             
newfilelocation = f'{queue_path}CA COG {file_report_date}.xlsx'

if os.path.exists(downloaded_file):            
    df_list = pd.read_html(downloaded_file)
    df = pd.DataFrame(df_list[0])          
    df.to_excel(newfilelocation,index=False)     
    os.remove(downloaded_file)        
    driver.quit()


    workbook = openpyxl.load_workbook(newfilelocation)
    workbook["Sheet1"].title = "CA COG"
    ws = workbook["CA COG"]
    workbook["CA COG"].delete_rows(ws.min_row, 1)
    for row in range(2, ws.max_row+1):
        ws["{}{}".format("J", row)].number_format = '@'  
        
    # Save the changes
    workbook.save(newfilelocation)

    print('Tapos Na!!!!')

 